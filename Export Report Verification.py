import os
import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

# Download folder
download_path = os.path.join(os.getcwd(), "Downloads")

if not os.path.exists(download_path):
    os.makedirs(download_path)

# Chrome Options
options = webdriver.ChromeOptions()

prefs = {
    "download.default_directory": download_path,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True
}

options.add_experimental_option("prefs", prefs)

driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=options
)

driver.maximize_window()

wait = WebDriverWait(driver, 10)

try:

    # Open Application
    driver.get("https://example.com/student-report")

    # -----------------------------
    # Apply Filter
    # -----------------------------
    wait.until(
        EC.presence_of_element_located((By.ID, "department"))
    ).send_keys("Computer Science")

    driver.find_element(By.ID, "searchButton").click()

    time.sleep(3)

    # -----------------------------
    # Count UI Rows
    # -----------------------------
    rows = driver.find_elements(By.XPATH, "//table/tbody/tr")

    ui_count = len(rows)

    print("Rows in UI :", ui_count)

    # -----------------------------
    # Export Excel
    # -----------------------------
    driver.find_element(By.ID, "exportExcel").click()

    print("Downloading Excel...")

    time.sleep(8)

    # -----------------------------
    # Verify Download
    # -----------------------------
    files = os.listdir(download_path)

    excel_files = [file for file in files if file.endswith(".xlsx")]

    assert len(excel_files) > 0, "Excel file not downloaded"

    latest_file = max(
        [os.path.join(download_path, f) for f in excel_files],
        key=os.path.getctime
    )

    print("Downloaded File :", latest_file)

    # -----------------------------
    # Read Excel
    # -----------------------------
    workbook = load_workbook(latest_file)

    sheet = workbook.active

    excel_rows = sheet.max_row - 1

    print("Rows in Excel :", excel_rows)

    # -----------------------------
    # Validation
    # -----------------------------
    assert ui_count == excel_rows

    print("PASS : UI data matches Excel data.")

except Exception as e:

    print("Test Failed")
    print(e)

finally:

    time.sleep(5)

    driver.quit()